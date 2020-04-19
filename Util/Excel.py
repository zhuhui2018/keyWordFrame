#encoding:utf-8

import openpyxl
from openpyxl import load_workbook
import os
import traceback

class ParseExcel():
    def __init__(self,excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = load_workbook(excel_file_path)
        self.sheet = self.set_sheet_by_name(self.wb.sheetnames[0])

    def get_excel_file_path(self):
        return self.excel_file_path

    def set_sheet_by_name(self,sheet_name=None):
        if sheet_name in self.wb.sheetnames:
            self.sheet = self.wb[sheet_name]
            return self.sheet
        self.sheet = None
        return self.sheet

    def get_sheet_by_name(self,name):
        return self.wb[name]

    def get_all_sheet_objects(self):
        sheets = []
        for sheetname in self.wb.sheetnames:
            sheets.append(self.get_sheet_by_name(sheetname))
        return sheets

    def get_all_sheet_names(self):
        return self.wb.sheetnames

    def get_current_sheet_name(self):
        return self.sheet.title

    def get_min_row(self):
        try:
            return self.sheet.min_row
        except:
            traceback.print_exc()

    def get_max_row(self):
        try:
            return self.sheet.max_row
        except:
            traceback.print_exc()
        
    def get_min_col(self):
        try:
            return self.sheet.min_column
        except:
            traceback.print_exc()

    def get_max_col(self):
        try:
            return self.sheet.max_column
        except:
            traceback.print_exc()

    def get_row(self,row_no):
        if not isinstance(row_no,int):
            return None
        try:
            return list(self.sheet.rows)[row_no-1]
        except:
            traceback.print_exc()

    def get_col(self,col_no):
        if not isinstance(col_no,int):
            return None
        try:
            return list(self.sheet.columns)[col_no - 1]
        except:
            traceback.print_exc()

    def write_cell(self,row_no,col_no,content):
        #if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            #return None
        try:
            self.sheet.cell(row=row_no,column=col_no).value = content
            self.wb.save(self.excel_file_path)
        except:
            traceback.print_exc()

    def get_cell_value(self,row_no,col_no):
        if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            return None
        try:
            return self.sheet.cell(row_no,col_no).value
        except:
            traceback.print_exc()


if __name__ == "__main__":
    file_path=r"D:\keyword_drivenn\testdata\测试用例.xlsx"
    wb = ParseExcel(file_path)
    print(wb.set_sheet_by_name("测试用例"))
    print(wb.get_current_sheet_name())
    print(wb.get_max_col())
    print(wb.get_min_row())
    print(wb.get_cell_value(2,1))
    wb.write_cell(1,1,"www.126.com")
    print(wb.get_sheet_by_name("测试用例"))
    print(wb.get_all_sheet_objects())
    print(wb.get_all_sheet_names())


